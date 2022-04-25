# -*- coding: utf-8 -*- 
"""
Created on 2021/11/4 11:49 
@File  : beauty_excel.py
@author: zhoul
@Desc  :
"""

import openpyxl
import os
from openpyxl.styles import Font, Color  # 设置字体、aRGB颜色
from openpyxl.styles import PatternFill, colors  # 设置背景色、颜色
from openpyxl.styles import Border, Side  # 设置边框
from openpyxl.styles import Alignment  # 设置单元格文字对齐方式、自动换行

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]


def beauty_format(file):
    wb = openpyxl.load_workbook(file)
    for i in range(len(wb.sheetnames)):
        ws = wb.worksheets[i]
        for irow, row in enumerate(ws.rows, start=1):
            if irow == 1:
                # 表头加粗、黑体
                color1 = PatternFill(fill_type="solid", fgColor="FFBCD2EE")
                font = Font('黑体', bold=True, size=14)
                for cell in row:
                    cell.fill = color1
                    cell.font = font
            elif irow % 2 == 0:
                # 偶数行
                font = Font('宋体', color='000000')
                for cell in row:
                    cell.font = font
            else:
                # 奇数行浅蓝色，宋体
                color1 = PatternFill("solid", fgColor="FFE5E5E5")
                font = Font('宋体', color='000000')
                for cell in row:
                    cell.fill = color1
                    cell.font = font
            ws.row_dimensions[1].heigth = 50  # 设置第一行的行高50
            ws.column_dimensions['B'].width = 25  # 设置B列的列宽25
            ws.column_dimensions['C'].width = 15  # 设置C列的列宽25
            ws.column_dimensions['D'].width = 16  # 设置D列的列宽25
            ws.column_dimensions['E'].width = 20  # 设置E列的列宽20
            ws.column_dimensions['F'].width = 60  # 设置F列的列宽50
            ws.column_dimensions['G'].width = 87  # 设置G列的列宽87
        # 另存为新文
        wb.save(file)


# beauty_format(rootPath + '\\testresults\\resultfile\\2021-11-04_测试环境_validation_error_result.xlsx')
